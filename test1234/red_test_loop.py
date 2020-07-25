import json
import unittest
import openpyxl
import urllib3
from config.PATH import *
import ddt
from test1234.red_test_do_excel import DoExcel
from core.send_requests import BaseCse
<<<<<<< HEAD
from core.my_logger import CaseLogModule
from core.my_logger import MyLog
=======
from core.my_log import CaseLogModule
from core.my_log import MyLog
>>>>>>> origin/master
from utils.my_functions import *
from utils.get_case_data import GetCaseData
from utils.common import *

ll = MyLog(get_module_name())
log_me456 = ll.get_logger()

new_case_data = GetCaseData().get_case_data()


@ddt.ddt
class TestApiDemo(BaseCse):

    def setUp(self):

        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        log_me456 .info('')
        log_me456 .info('----------------------  用例开始  ----------------------')

    @ddt.data(*new_case_data)
    def test_api_case(self, data1):
        # 分割用例数据
        global test_result
        if data1['header'] != "":
            header = json.loads(json.dumps(eval(data1['header'])))
        else:
            header = None
        case_name = data1['case_name']
        case_id = data1['case_id']
        url = data1['url']
        if isinstance(data1['method'], dict):
            method = json.dumps(data1['method'])
        else:
            method = data1['method']

        if isinstance(data1['data'], dict) or isinstance(data1['data'], list) or isinstance(data1['data'], tuple):
            data = json.dumps(data1['data'])
        else:
            data = data1['data']

        if 'my_' in str(data):
            new_data = json.dumps(eval(data))
        else:
            new_data = data

        if data1['files'] == "":
            files = None
        elif data1['files'] != "" and isinstance(data1['files'], str):
            files = eval(data1['files'])
        else:
            files = data1['files']

        msg = data1['msg']

        if isinstance(data1['proceed'], dict) or isinstance(data1['proceed'], list) or isinstance(data1['proceed'], tuple):
            proceed = json.dumps(eval(data1['proceed']))
        else:
            proceed = data1['proceed']

        # 打印日志
        CaseLogModule.case_log_module_request(case_id, case_name, url, method, header, new_data)

        # 发送请求
        response = json.loads(self.send_requests(url, method, header, new_data, files))

        # 打印日志
        CaseLogModule.case_log_module_response(msg, response, proceed)

        # 添加变量，为下一个接口做准备
        if proceed != "":
            log_me456 .info('添加全局变量：{}'.format(proceed))
            eval(proceed)
        else:
            pass

        try:
            # 断言
            self.assertIn(msg, json.dumps(response, ensure_ascii=False))
            test_result = 'PASS'
        except AssertionError as e:
            test_result = 'FAIL'
            log_me456.error(e)
            raise e
        finally:
            log_me456.info('-----------------------------> 对Excel写入测试结果 <-----------------------------')
            base_case_filename = os.path.join(BASE_CASE_DATA, 'api_case.xlsx')
            # 统计csv的数量，来决定 被输入行数 是否需要再增加
            if new_case_data.index(data1) == 0:
                my_setattr('csv_count', 0)
            elif data1['csv_loop'] != my_getattr('now_csv_name'):
                my_setattr('csv_count', my_getattr('csv_count') + 1)

            # 初始化 - 被输入的行数（列数是固定的位置，不用初始化）
            # 只要文件名、表名任何一个名字不再相同，就会进行 初始化
            if new_case_data.index(data1) == 0 or \
                    data1['sheet_name'] != my_getattr('now_table_name') or \
                    data1['csv_loop'] != my_getattr('now_csv_name'):
                my_setattr('base_row_n', 2)
                my_setattr('csv_row_n', 2)
                my_setattr('now_table_name', data1['sheet_name'])
                my_setattr('now_csv_name', data1['csv_loop'])
            else:
                my_setattr('now_table_name', data1['sheet_name'])
                my_setattr('now_csv_name', data1['csv_loop'])

            if data1['csv_loop'] != "":  # 对csv用例文件写入结果
                csv_result_filepath = os.path.join(CSV_CASE_DATA, data1['csv_loop'].split('/')[0])
                csv_result_sheet_name = data1['csv_loop'].split('/')[1]
                excel = openpyxl.load_workbook(csv_result_filepath)
                table = excel[csv_result_sheet_name]
                max_col = table.max_column
                # 写入结果
                DoExcel(csv_result_filepath, csv_result_sheet_name).write_excel(
                                                         test_result,
                                                         my_getattr('csv_row_n'),
                                                         max_col-2)
                # 写入测试时间
                DoExcel(csv_result_filepath, csv_result_sheet_name).write_excel(
                                                         my_now_time(),
                                                         my_getattr('csv_row_n'),
                                                         max_col)
                my_setattr('csv_row_n', my_getattr('csv_row_n')+1)  # 加一行
            else:  # 对base用例文件写入结果
                excel = openpyxl.load_workbook(base_case_filename)
                table = excel[data1['sheet_name']]
                max_col = table.max_column
                # 写入结果
                DoExcel(base_case_filename, data1['sheet_name']).write_excel(
                                                        test_result,
                                                        my_getattr('csv_count') + my_getattr('base_row_n'),
                                                        max_col-3)
                # 写入测试时间
                DoExcel(base_case_filename, data1['sheet_name']).write_excel(
                                                        my_now_time(),
                                                        my_getattr('csv_count') + my_getattr('base_row_n'),
                                                        max_col)
                my_setattr('base_row_n', my_getattr('base_row_n')+1)  # 加一行

    def tearDown(self):
        log_me456 .info('----------------------  用例结束  ----------------------')
        log_me456 .info('')


if __name__ == '__main__':
    unittest.main()
