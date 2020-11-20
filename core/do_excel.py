import re
import openpyxl
from openpyxl.styles import Font, Alignment, colors
from core.my_logger import log
import json
from utils.common import *


class DoExcel:
    def __init__(self, filepath, sheet_name):
        self.filepath = filepath
        self.sheet_name = sheet_name
        log.info('打开 {} 的 {} 表'.format(self.filepath, self.sheet_name))
        self.excel = openpyxl.load_workbook(self.filepath)  # 打开Excel
        self.table = self.excel[self.sheet_name]

    # def __del__(self):
    #     self.excel.close()  # 一定要关闭Excel

    def read_excel(self, case_order=None):
        case_data = []
        table = self.table
        max_row = table.max_row  # 最大行
        max_col = table.max_column  # 最大列

        #  获取首行的  case_title
        case_title = []
        for i in range(1, max_col + 1):  # 获取标题行
            cell_value = table.cell(row=1, column=i).value
            case_title.append(cell_value)

        # 设置 获取哪些行的数据
        if case_order == '' or case_order is None or case_order == 'null':
            case_order = 'all'  # 默认取  all
        else:
            case_order = case_order

        case_value = []  # 每一行的数据
        if isinstance(case_order, str) and case_order.lower() == 'all':  # 获取所有行的数据
            log.info('读取 {} 表 - 全部行的 - 用例数据'.format(self.sheet_name))
            for row in range(max_row):  # 获取每一行的值
                if row+1 == max_row:  # 超出了最大行就不再去取值
                    break
                for col in range(max_col):  # 获取一整行的值
                    value = table.cell(row+2, col+1).value  # 获取每一个单元格的value

                    if value is None or value == "" or value == '':
                        case_value.append("")
                    else:
                        case_value.append(value)

                # 看看一整行数据是否为空
                b_str = ''
                for a_idx, a_value in enumerate(case_value):
                    b_str += str(a_value)
                if re.sub(r'\s+', "", b_str) != "":
                    case_data.append(dict(zip(case_title, case_value)))  # 压缩进 case_data
                    case_value.clear()  # 清空，下次循环才能正常进行
                elif re.sub(r'\s+', "", b_str) == "":
                    case_value.clear()
                    log.warning('{0} 表 {1} 整行为空，这一整行不纳入获取结果，请知悉。'.format(self.sheet_name, row+2))
                    continue
        elif isinstance(case_order, list):  # 获取部分行的数据
            for row2 in case_order:
                if row2 == max_row:  # 超出了最大行就不再取值了
                    break
                elif isinstance(row2, str):  # 如果是str，就转成int
                    row2 = int(row2)
                log.info('读取 {} 表 - 第{}行的 - 用例数据'.format(self.sheet_name, row2))
                for col2 in range(max_col):  # 获取一整行的值
                    value = table.cell(row2+1, col2+1).value  # 获取每一个单元格的value

                    if value is None or value == "" or value == '':
                        case_value.append("")
                    else:
                        case_value.append(value)

                # 看看一整行数据是否为空
                b_str = ''
                for a_idx, a_value in enumerate(case_value):
                    b_str += str(a_value)
                if re.sub(r'\s+', "", b_str) != "":
                    case_data.append(dict(zip(case_title, case_value)))  # 压缩进 case_data
                    case_value.clear()
                elif re.sub(r'\s+', "", b_str) == "":
                    log.warning('{0} 表 {1} 整行为空，这一整行不纳入获取结果，请知悉。'.format(self.sheet_name, row2+1))
                    continue
        # log.info('读取Excel - 反值：{}'.format(json.dumps(case_data, indent=2, ensure_ascii=False)))
        return case_data

    def write_excel(self, value, row_n, col_n=None):
        """
        :param value: 内容
        :param row_n:第几行
        :param col_n: 第几列
        """
        try:
            if col_n is None:
                col_n = 9

            table = self.table

            # 定义字体格式
            font_green = Font(name='宋体', color=colors.GREEN, bold=True)  # 绿色
            font_red = Font(name='宋体', color=colors.RED, bold=True)  # 红色
            # font_dark_yellow = Font(name='宋体', color=colors.DARKYELLOW, bold=True)  # 暗黄色
            # align = Alignment(horizontal='center', vertical='center')  # 居中
            align = Alignment(horizontal='left', vertical='center')  # 居左

            # 写入测试结果，带有格式
            if value.lower() == "pass":
                table.cell(row_n, col_n).value = value  # 写入pass
                table.cell(row_n, col_n).font = font_green  # 绿色
            elif value.lower() == "fail":
                table.cell(row_n, col_n).value = value  # 写入fail
                table.cell(row_n, col_n).font = font_red  # 红色
            else:
                table.cell(row_n, col_n).value = value  # 写入任一值

            # 字体居格式
            table.cell(row_n, col_n).alignment = align
            table.cell(row_n, col_n).alignment = align

            # 保存修改
            self.excel.save(self.filepath)
            log.info('修改{}成功：({},{})={}'.format(os.path.split(self.filepath)[1], row_n, col_n, value))
        except PermissionError:
            log.warning('修改{0}失败：请关闭已经打开的 {0}。'.format(os.path.split(self.filepath)[1]))
        except KeyError:
            log.warning('修改{}失败：没有 {} 表'.format(os.path.split(self.filepath)[1], self.sheet_name))


if __name__ == '__main__':
    # base_file = r'H:\AutoFramework_Open\AutoTest_API\test_data\base_case_data\api_case.xlsx'
    # case_data1 = DoExcel(base_file, 'login').read_excel()
    # case_data = json.dumps(case_data, ensure_ascii=False)
    # w.write_excel('home', 'PASS', 2)
    # w.write_excel('home', 'FAIL', 3)
    # print(type(case_data))
    # print(case_data)
    # print(json.dumps(case_data))

    # csv_file = r'H:\AutoTest\API\my_api_autotest\test_data\csv_case_data\login.xlsx'
    # csv_data = DoExcel(csv_file).read_excel('login')
    # print(csv_data)
    pass
