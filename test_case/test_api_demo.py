import json
import unittest
import openpyxl
import urllib3
from config.PATH import *
import ddt
from core.do_excel import DoExcel
from core.send_requests import BaseCse
from core.my_logger import log
from utils.my_functions import *
from core.get_case_data import GetCaseData
from utils.common import *
from utils.my_split import deal_function_data, get_str

# 获取原始用例数据 和 CSV的参数化用例数据，并且合并成最终的用例数据。
new_case_data = GetCaseData().get_case_data()
response = None
test_result = None


@ddt.ddt
class TestApiDemo(BaseCse):
    def setUp(self):
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        log.info('')
        log.info('<' * 30 + '  用例开始  ' + '>' * 30)

    @ddt.data(*new_case_data)
    def test_api_case(self, data1):
        # 分割用例数据
        global test_result
        global response
        case_id = data1['case_id']
        case_name = data1['case_name']
        url = data1['url']
        method = deal_function_data(data1['method'])

        # 请求header
        header = deal_function_data(data1['header'])

        # 请求form_data
        param = deal_function_data(data1['param'])
        if param == "" or param == '' or param is None or param == 'null':
            param = None
        else:
            param = eval(param)

        # 请求body_data
        body = deal_function_data(data1['body'])
        if body == "" or body == '' or body is None or body == 'null':
            body = None
        else:
            body = eval(body)

        # 请求上传文件
        files = deal_function_data(data1['files'])
        msg = deal_function_data(data1['msg'])
        proceed = get_str(data1['proceed'])

        # 打印日志
        log.info('用例ID：{}'.format(case_id))
        log.info('用例name：{}'.format(case_name))

        try:
            # 发送请求
            response = json.loads(self.send_requests(url, method, header, param, body, files))
            # 打印日志
            log.info(('期望结果：{}'.format(msg)))
            log.info(('实际结果：{}'.format(response)))

            # 添加变量，为下一个接口做准备
            if len(proceed) != 0:
                log.info('执行proceed：{}'.format(proceed))
                for i in proceed:
                    eval(i)

            # 断言
            self.assertIn(msg, json.dumps(response, ensure_ascii=False))
            test_result = 'PASS'
        except Exception as e:
            test_result = 'FAIL'
            self.fail(e)
            # raise Exception
        finally:  # 对每一条用例写入Excel的测试结果
            log.info("-" * 20 + "> 对Excel写入测试结果 <" + "-" * 20)
            base_case_filename = os.path.join(BASE_CASE_DATA, 'api_case.xlsx')
            identify = data1['sheet_name'] + data1['csv_loop']

            # 统计csv的数量，来决定 被输入行数 是否需要再增加
            if new_case_data.index(data1) == 0 and data1['csv_loop'] != "":
                my_setattr('csv_count', 1)
            elif new_case_data.index(data1) == 0:
                my_setattr('csv_count', 0)
            # 新的csv_name不能是"" + 老的csv_name不能是"" + 新的csv_name != 老的csv_name。这样才算是真正的多了一条参数化用例数据。
            elif data1['csv_loop'] != "" and my_getattr('csv_name') != "" and data1['csv_loop'] != my_getattr(
                    'csv_name'):
                my_setattr('csv_count', my_getattr('csv_count') + 1)

            # 初始化 - 被输入的行数（列数是固定的位置，不用初始化）
            # 只要文件名、表名任何一个名字不再相同，就会进行 初始化
            if new_case_data.index(data1) == 0 or identify != my_getattr('identify'):
                my_setattr('base_row_n', 2)
                my_setattr('csv_row_n', 2)
                my_setattr('identify', identify)
            else:
                my_setattr('identify', identify)

            if data1['csv_loop'] != "":  # 对csv用例文件写入结果
                csv_result_filepath = os.path.join(CSV_CASE_DATA, data1['csv_loop'].split('/')[0])
                csv_result_sheet_name = data1['csv_loop'].split('/')[1]
                excel = openpyxl.load_workbook(csv_result_filepath)
                table = excel[csv_result_sheet_name]
                max_col = table.max_column
                # 写入结果
                DoExcel(csv_result_filepath, csv_result_sheet_name).write_excel(
                    test_result,
                    my_getattr('csv_row_n'),
                    max_col - 2)
                # 写入测试时间
                DoExcel(csv_result_filepath, csv_result_sheet_name).write_excel(
                    my_now_time(),
                    my_getattr('csv_row_n'),
                    max_col)
                my_setattr('csv_row_n', my_getattr('csv_row_n') + 1)  # 加一行
            else:  # 对base用例文件写入结果
                excel = openpyxl.load_workbook(base_case_filename)
                table = excel[data1['sheet_name']]
                max_col = table.max_column
                # 写入结果
                DoExcel(base_case_filename, data1['sheet_name']).write_excel(
                    test_result,
                    my_getattr('csv_count') + my_getattr('base_row_n'),
                    max_col - 3)
                # 写入测试时间
                DoExcel(base_case_filename, data1['sheet_name']).write_excel(
                    my_now_time(),
                    my_getattr('csv_count') + my_getattr('base_row_n'),
                    max_col)
                my_setattr('base_row_n', my_getattr('base_row_n') + 1)  # 加一行

            # 老的csv_name：把csv_loop设置成全局变量，为下一条的  统计csv数量  做准备
            my_setattr('csv_name', data1['csv_loop'])

    def tearDown(self):
        log.info('<' * 30 + '  用例结束  ' + '>' * 30)
        log.info('')


if __name__ == '__main__':
    unittest.main()
